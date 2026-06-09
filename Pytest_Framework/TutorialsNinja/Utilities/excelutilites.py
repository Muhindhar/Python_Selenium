import openpyxl
def get_data(path,sheetname):
    final_list=[]
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    tot_row = sheet.max_row
    tot_col = sheet.max_column
    
    for i in range(2,tot_row+1):
        row_list=[]
        for c in range(1,tot_col+1):
            row_list.append(sheet.cell(i,c).value)
        final_list.append(row_list)
    return final_list